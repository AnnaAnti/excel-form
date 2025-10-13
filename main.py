import os
from math import floor, ceil
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from openpyxl import load_workbook

# === ПУТИ И НАСТРОЙКИ ===
BASE_DIR = os.path.dirname(__file__)
EXCEL_FILE = os.getenv("EXCEL_FILE", os.path.join(BASE_DIR, "book.xlsm"))  # шаблон
REPORTS_DIR = os.getenv("REPORTS_DIR", os.path.join(BASE_DIR, "reports"))
SHEET1_NAME = os.getenv("SHEET1_NAME", "Лист1")
SHEET2_NAME = os.getenv("SHEET2_NAME", "Лист2")

os.makedirs(REPORTS_DIR, exist_ok=True)

app = FastAPI(title="Excel Sheet2 Writer")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# Раздаём статические файлы: корень = index.html, отчёты = /reports/...
app.mount("/reports", StaticFiles(directory=REPORTS_DIR), name="reports")

@app.get("/")
def root():
    index_path = os.path.join(BASE_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(404, "index.html not found")
    return FileResponse(index_path, media_type="text/html")

# === Схемы ===
class Guest(BaseModel):
    fio: str = ""
    position: str = ""

class FormPayload(BaseModel):
    card_owner: str
    assistant: str
    restaurant: str
    address: str

    check1_number: str
    check1_date: str
    check1_time_decimal: float
    check1_sum: float

    check2_number: Optional[str] = ""
    check2_date: Optional[str] = ""
    check2_time_decimal: Optional[float] = None
    check2_sum: Optional[float] = None

    company: Optional[str] = ""
    goal: Optional[str] = ""
    topic: Optional[str] = ""
    totals: Optional[str] = ""

    guests: List[Guest] = Field(default_factory=list)

# === Вспомогательные ===
def money_to_words(amount: float) -> str:
    # Упрощённо, можно оставить вашу расширенную версию — суть не меняется
    def plural(n, f1, f2, f5):
        n = abs(int(n)) % 100
        n1 = n % 10
        if 11 <= n <= 19: return f5
        if n1 == 1: return f1
        if 2 <= n1 <= 4: return f2
        return f5
    rub = int(amount)
    kop = int(round((amount - rub) * 100))
    return f"{rub} {plural(rub,'рубль','рубля','рублей')} {kop:02d} {plural(kop,'копейка','копейки','копеек')}"

def _num(v) -> float:
    try: return float(v)
    except: return 0.0

# === Эндпоинты ===
@app.get("/api/options")
def get_options():
    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(404, f"Excel template not found: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, keep_vba=True, data_only=True)
    ws1 = wb[SHEET1_NAME]

    owners = [str(ws1[f"A{r}"].value) for r in range(2, 7) if ws1[f"A{r}"].value not in (None, "")]
    assistants = [str(ws1[f"A{r}"].value) for r in range(8, 14) if ws1[f"A{r}"].value not in (None, "")]
    return {"owners": owners, "assistants": assistants}

@app.post("/api/submit")
def submit_form(p: FormPayload):
    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(404, f"Excel template not found: {EXCEL_FILE}")

    # 1) Открываем шаблон, пишем ТОЛЬКО лист2
    wb = load_workbook(EXCEL_FILE, keep_vba=True, data_only=False)
    ws2 = wb[SHEET2_NAME]

    ws2["B5"].value = p.card_owner
    ws2["B6"].value = p.assistant
    ws2["B9"].value = p.restaurant
    ws2["B17"].value = p.address

    ws2["C9"].value = p.check1_number
    ws2["D9"].value = p.check1_date
    ws2["E9"].value = p.check1_time_decimal
    ws2["A9"].value = p.check1_sum

    if p.check2_number: ws2["C10"].value = p.check2_number
    if p.check2_date:   ws2["D10"].value = p.check2_date
    if p.check2_time_decimal is not None: ws2["E10"].value = p.check2_time_decimal
    if p.check2_sum is not None:          ws2["A10"].value = p.check2_sum

    if p.company: ws2["B33"].value = p.company
    if p.goal:    ws2["B34"].value = p.goal
    if p.topic:   ws2["C37"].value = p.topic
    if p.totals:  ws2["A44"].value = p.totals

    base_row = 37
    for i in range(5):
        fio = p.guests[i].fio if i < len(p.guests) else ""
        pos = p.guests[i].position if i < len(p.guests) else ""
        ws2[f"A{base_row+i}"].value = fio
        ws2[f"B{base_row+i}"].value = pos

    # 2) Считаем A14 и C26 на сервере и пропись
    a_vals = [ws2[f"A{r}"].value for r in range(9, 14)]
    a14_calc = sum(_num(v) for v in a_vals)
    c26_calc = 1000 * ceil(a14_calc / 1000.0) if a14_calc != 0 else 0.0
    ws2["B15"].value = money_to_words(a14_calc)
    ws2["B27"].value = money_to_words(c26_calc)

    # 3) Сохраняем НОВУЮ КОПИЮ в reports/
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"book_{ts}.xlsm"
    out_path = os.path.join(REPORTS_DIR, out_name)
    wb.save(out_path)

    # Возвращаем ссылку на скачивание
    return {
        "status": "ok",
        "a14": a14_calc,
        "c26": c26_calc,
        "download_url": f"/reports/{out_name}"
    }
